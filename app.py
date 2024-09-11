import Adafruit_DHT
from openpyxl import load_workbook
import time

# Configurações do sensor
sensor = Adafruit_DHT.DHT11
pino_sensor = 4  # GPIO que conecta o sensor DHT11 (ajuste conforme necessário)

# Caminho para o arquivo Excel
arquivo_excel = 'dados_sensores.xlsx'

# Função para ler o sensor de temperatura e umidade
def ler_dados_sensor():
    umidade, temperatura = Adafruit_DHT.read_retry(sensor, pino_sensor)
    if umidade is not None and temperatura is not None:
        return temperatura, umidade
    else:
        return None, None

# Função para escrever os dados na planilha Excel
def escrever_no_excel(temperatura, umidade):
    # Carregar o arquivo Excel existente
    workbook = load_workbook(arquivo_excel)
    planilha = workbook.active

    # Encontrar a próxima linha vazia
    linha = planilha.max_row + 1

    # Inserir os dados na planilha
    planilha.cell(row=linha, column=1).value = time.strftime("%Y-%m-%d %H:%M:%S")  # Data e hora
    planilha.cell(row=linha, column=2).value = temperatura  # Temperatura
    planilha.cell(row=linha, column=3).value = umidade  # Umidade

    # Salvar as alterações
    workbook.save(arquivo_excel)
    print(f'Dados gravados: Temperatura = {temperatura}ºC, Umidade = {umidade}%')

# Loop principal para ler os dados periodicamente
while True:
    temperatura, umidade = ler_dados_sensor()
    if temperatura is not None and umidade is not None:
        escrever_no_excel(temperatura, umidade)
    else:
        print('Falha na leitura do sensor. Tentando novamente...')
    
    # Esperar 10 minutos antes da próxima leitura (ajuste conforme necessário)
    time.sleep(600)
